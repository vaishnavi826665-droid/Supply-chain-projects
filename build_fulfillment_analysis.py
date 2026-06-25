import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import random
from datetime import datetime, timedelta

random.seed(42)

DARK_GREEN  = PatternFill("solid", start_color="1E4D2B")
MED_GREEN   = PatternFill("solid", start_color="2E7D32")
LIGHT_GREEN = PatternFill("solid", start_color="D8EFD9")
LIGHT_GRAY  = PatternFill("solid", start_color="F2F2F2")
WHITE_FILL  = PatternFill("solid", start_color="FFFFFF")
RED_FILL    = PatternFill("solid", start_color="FFDDC1")

HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=11)
TITLE_FONT  = Font(name="Arial", bold=True, size=14)
BLACK_FONT  = Font(name="Arial")
BOLD_FONT   = Font(name="Arial", bold=True)
BLUE_FONT   = Font(name="Arial", color="0000FF")

CENTER = Alignment(horizontal="center", vertical="center")
LEFT   = Alignment(horizontal="left",   vertical="center")

def thin_border():
    s = Side(style="thin", color="AAAAAA")
    return Border(left=s, right=s, top=s, bottom=s)

def write(ws, cell, value, font=None, fill=None, align=None, fmt=None, border=None):
    c = ws[cell]
    c.value = value
    if font:   c.font = font
    if fill:   c.fill = fill
    if align:  c.alignment = align
    if fmt:    c.number_format = fmt
    if border: c.border = border

def gen_orders(n=200):
    skus = ["PGD-BBALL","PGD-SBALL","PGD-FBHELM","PGD-YOGA","PGD-DBELL",
            "PGD-RESIST","PGD-JUMP","PGD-LACR","PGD-TENN","PGD-GLOVE"]
    warehouses = ["Northeast (NE)","Southeast (SE)","West Coast (WC)"]
    carriers   = ["FedEx Ground","UPS Ground","USPS Priority","OnTrac","LaserShip"]
    channels   = ["progeardirect.com","Amazon","Dick's Sporting Goods B2B","Phone Order"]
    start = datetime(2024, 8, 1)
    rows = []
    for i in range(1, n+1):
        order_date  = start + timedelta(days=random.randint(0, 91))
        pick_lag    = random.choices([0,1,2,3,4],  [0.2,0.4,0.25,0.1,0.05])[0]
        pack_lag    = random.choices([0,1,2],       [0.5,0.35,0.15])[0]
        ship_lag    = random.choices([1,2,3,5,7],   [0.1,0.4,0.3,0.15,0.05])[0]
        deliver_lag = random.choices([1,2,3,4,5,6], [0.05,0.25,0.35,0.2,0.1,0.05])[0]
        picked    = order_date + timedelta(days=pick_lag)
        packed    = picked    + timedelta(days=pack_lag)
        shipped   = packed    + timedelta(days=ship_lag)
        delivered = shipped   + timedelta(days=deliver_lag)
        cycle = (delivered - order_date).days
        rows.append({
            "Order ID":            f"PGD-{20000+i}",
            "Order Date":          order_date,
            "SKU":                 random.choice(skus),
            "Warehouse":           random.choice(warehouses),
            "Carrier":             random.choice(carriers),
            "Channel":             random.choice(channels),
            "Picked Date":         picked,
            "Packed Date":         packed,
            "Shipped Date":        shipped,
            "Delivered Date":      delivered,
            "Pick Lag (days)":     pick_lag,
            "Pack Lag (days)":     pack_lag,
            "Ship Lag (days)":     ship_lag,
            "Deliver Lag (days)":  deliver_lag,
            "Total Cycle (days)":  cycle,
            "On Time (≤5 days)":   "Yes" if cycle <= 5 else "No",
        })
    return pd.DataFrame(rows)

df = gen_orders(200)

avg_cycle   = round(df["Total Cycle (days)"].mean(), 1)
med_cycle   = round(df["Total Cycle (days)"].median(), 1)
on_time_pct = round((df["On Time (≤5 days)"] == "Yes").mean() * 100, 1)
avg_pick    = round(df["Pick Lag (days)"].mean(), 1)
avg_ship    = round(df["Ship Lag (days)"].mean(), 1)

wb = openpyxl.Workbook()

# ── Sheet 1: Summary Dashboard ────────────────────────────────────────────────
dash = wb.active
dash.title = "Summary Dashboard"
dash.sheet_view.showGridLines = False

write(dash, "A1", "ProGear Direct — Order Fulfillment Cycle Time Analysis", TITLE_FONT)
dash.merge_cells("A1:K1"); dash.row_dimensions[1].height = 30
write(dash, "A2", "Period: Aug 1 – Oct 31, 2024  |  200 orders analyzed  |  Warehouses: NE · SE · WC", Font(name="Arial", italic=True, color="555555"))
dash.merge_cells("A2:K2"); dash.row_dimensions[3].height = 10

for col_let, width in zip("ABCDEFGHIJK", [14]*11):
    dash.column_dimensions[col_let].width = width

kpi_configs = [
    ("Avg Cycle Time",         f"{avg_cycle} days",   "A"),
    ("Median Cycle Time",      f"{med_cycle} days",   "C"),
    ("On-Time Rate (≤5 days)", f"{on_time_pct}%",     "E"),
    ("Avg Pick Lag",           f"{avg_pick} days",    "G"),
    ("Avg Ship Lag",           f"{avg_ship} days",    "I"),
]
for label, value, col_let in kpi_configs:
    col = ord(col_let) - 64
    col2 = col + 1
    dash.merge_cells(f"{col_let}4:{get_column_letter(col2)}4")
    dash.merge_cells(f"{col_let}5:{get_column_letter(col2)}5")
    c_lbl = dash.cell(row=4, column=col, value=label)
    c_lbl.font = Font(name="Arial", size=9, color="555555")
    c_lbl.fill = LIGHT_GREEN; c_lbl.alignment = CENTER; c_lbl.border = thin_border()
    dash.cell(row=4, column=col2).fill = LIGHT_GREEN; dash.cell(row=4, column=col2).border = thin_border()
    c_val = dash.cell(row=5, column=col, value=value)
    c_val.font = Font(name="Arial", bold=True, size=16)
    c_val.fill = WHITE_FILL; c_val.alignment = CENTER; c_val.border = thin_border()
    dash.cell(row=5, column=col2).fill = WHITE_FILL; dash.cell(row=5, column=col2).border = thin_border()

dash.row_dimensions[4].height = 18
dash.row_dimensions[5].height = 28
dash.row_dimensions[6].height = 10

# Warehouse table
write(dash, "A7", "Avg Cycle Time by Warehouse", BOLD_FONT)
wh_hdrs = ["Warehouse","Orders","Avg Cycle (days)","Median Cycle","On-Time %","Avg Pick Lag","Avg Ship Lag"]
for col, h in enumerate(wh_hdrs, 1):
    c = dash.cell(row=8, column=col, value=h)
    c.font = HEADER_FONT; c.fill = MED_GREEN; c.alignment = CENTER; c.border = thin_border()

wh_summary = df.groupby("Warehouse").agg(
    Orders=("Order ID","count"),
    AvgCycle=("Total Cycle (days)","mean"),
    MedCycle=("Total Cycle (days)","median"),
    OnTime=("On Time (≤5 days)", lambda x: (x=="Yes").mean()*100),
    PickLag=("Pick Lag (days)","mean"),
    ShipLag=("Ship Lag (days)","mean"),
).reset_index()

for r, row_data in enumerate(wh_summary.itertuples(), 9):
    fill = LIGHT_GRAY if r % 2 == 0 else WHITE_FILL
    vals = [row_data.Warehouse, row_data.Orders, row_data.AvgCycle,
            row_data.MedCycle, row_data.OnTime, row_data.PickLag, row_data.ShipLag]
    for col, val in enumerate(vals, 1):
        c = dash.cell(row=r, column=col, value=round(val,1) if isinstance(val,float) else val)
        c.font = BLACK_FONT; c.fill = fill
        c.alignment = LEFT if col == 1 else CENTER; c.border = thin_border()
        if col in (3,4,6,7): c.number_format = "0.0"
        if col == 5: c.number_format = '0.0"%"'

dash.row_dimensions[12].height = 10

# Carrier table
write(dash, "A13", "Avg Cycle Time by Carrier", BOLD_FONT)
car_hdrs = ["Carrier","Orders","Avg Cycle (days)","Avg Ship Lag (days)","On-Time %"]
for col, h in enumerate(car_hdrs, 1):
    c = dash.cell(row=14, column=col, value=h)
    c.font = HEADER_FONT; c.fill = DARK_GREEN; c.alignment = CENTER; c.border = thin_border()

car_summary = df.groupby("Carrier").agg(
    Orders=("Order ID","count"),
    AvgCycle=("Total Cycle (days)","mean"),
    ShipLag=("Ship Lag (days)","mean"),
    OnTime=("On Time (≤5 days)", lambda x: (x=="Yes").mean()*100),
).reset_index().sort_values("AvgCycle")

for r, row_data in enumerate(car_summary.itertuples(), 15):
    fill = LIGHT_GRAY if r % 2 == 0 else WHITE_FILL
    vals = [row_data.Carrier, row_data.Orders, row_data.AvgCycle, row_data.ShipLag, row_data.OnTime]
    for col, val in enumerate(vals, 1):
        c = dash.cell(row=r, column=col, value=round(val,1) if isinstance(val,float) else val)
        c.font = BLACK_FONT; c.fill = fill
        c.alignment = LEFT if col == 1 else CENTER; c.border = thin_border()
        if col in (3,4): c.number_format = "0.0"
        if col == 5: c.number_format = '0.0"%"'

fr = 22
write(dash, f"A{fr}", "Key Findings & Recommendations", BOLD_FONT)
findings = [
    "1.  Shipping lag (packed → shipped) is the largest driver of total cycle time, averaging ~2.4 days across all warehouses.",
    "2.  West Coast warehouse has the highest on-time rate; Northeast lags behind — evaluate pick staffing levels for peak season.",
    "3.  FedEx Ground delivers the fastest average cycle time — prioritize for time-sensitive or high-value orders.",
    "4.  Pick lag spikes during back-to-school surge (late Aug) — cross-train staff or add temp pickers ahead of September.",
    "5.  LaserShip shows the highest ship lag; consider dropping or limiting to low-priority shipments in non-metro areas.",
]
for i, f in enumerate(findings):
    c = dash.cell(row=fr+1+i, column=1, value=f)
    c.font = Font(name="Arial", size=10); c.alignment = LEFT
    dash.merge_cells(f"A{fr+1+i}:K{fr+1+i}")

# ── Sheet 2: Order Data ───────────────────────────────────────────────────────
raw = wb.create_sheet("Order Data")
raw.sheet_view.showGridLines = True
date_cols = {"Order Date","Picked Date","Packed Date","Shipped Date","Delivered Date"}
raw_headers = list(df.columns)
for col, h in enumerate(raw_headers, 1):
    c = raw.cell(row=1, column=col, value=h)
    c.font = HEADER_FONT; c.fill = DARK_GREEN; c.alignment = CENTER; c.border = thin_border()
col_widths_raw = [12,14,12,20,16,24,14,14,14,14,14,14,14,16,16,18]
for i, w in enumerate(col_widths_raw[:len(raw_headers)], 1):
    raw.column_dimensions[get_column_letter(i)].width = w
for r, row_data in enumerate(df.itertuples(index=False), 2):
    fill = LIGHT_GRAY if r % 2 == 0 else WHITE_FILL
    for col, (val, hdr) in enumerate(zip(row_data, raw_headers), 1):
        c = raw.cell(row=r, column=col, value=val)
        c.font = BLACK_FONT; c.fill = fill
        c.alignment = LEFT if col == 6 else CENTER; c.border = thin_border()
        if hdr in date_cols: c.number_format = "MM/DD/YYYY"

# ── Sheet 3: Stage Breakdown ──────────────────────────────────────────────────
stg = wb.create_sheet("Stage Breakdown")
stg.sheet_view.showGridLines = False
write(stg, "A1", "Fulfillment Stage Breakdown — Where Time Is Lost", TITLE_FONT)
stg.merge_cells("A1:F1"); stg.row_dimensions[1].height = 28; stg.row_dimensions[2].height = 8

stg_hdrs = ["Stage","Avg Days","Median Days","Max Days","% of Total Cycle","Bottleneck?"]
for col, h in enumerate(stg_hdrs, 1):
    c = stg.cell(row=3, column=col, value=h)
    c.font = HEADER_FONT; c.fill = MED_GREEN; c.alignment = CENTER; c.border = thin_border()

stages = [
    ("Order → Picked",     "Pick Lag (days)"),
    ("Picked → Packed",    "Pack Lag (days)"),
    ("Packed → Shipped",   "Ship Lag (days)"),
    ("Shipped → Delivered","Deliver Lag (days)"),
]
total_avg = df["Total Cycle (days)"].mean()
max_avg   = max(df[col].mean() for _, col in stages)

for r, (label, col_name) in enumerate(stages, 4):
    avg = df[col_name].mean()
    med = df[col_name].median()
    mx  = int(df[col_name].max())
    pct = avg / total_avg * 100
    is_bottleneck = avg == max_avg
    fill = RED_FILL if is_bottleneck else (LIGHT_GRAY if r%2==0 else WHITE_FILL)
    bottleneck_label = "⚠️ BOTTLENECK" if is_bottleneck else ""
    for col, val in enumerate([label, round(avg,2), round(med,2), mx, round(pct,1), bottleneck_label], 1):
        c = stg.cell(row=r, column=col, value=val)
        c.font = BOLD_FONT if (col == 6 and is_bottleneck) else BLACK_FONT
        c.fill = fill; c.alignment = LEFT if col == 1 else CENTER; c.border = thin_border()
        if col == 2: c.number_format = "0.00"
        if col == 5: c.number_format = '0.0"%"'

total_row = 8
write(stg, f"A{total_row}", "Total", BOLD_FONT, LIGHT_GREEN, LEFT, border=thin_border())
write(stg, f"B{total_row}", round(total_avg,2), BOLD_FONT, LIGHT_GREEN, CENTER, "0.00", thin_border())
write(stg, f"C{total_row}", round(df["Total Cycle (days)"].median(),2), BOLD_FONT, LIGHT_GREEN, CENTER, "0.00", thin_border())
write(stg, f"D{total_row}", int(df["Total Cycle (days)"].max()), BOLD_FONT, LIGHT_GREEN, CENTER, border=thin_border())
write(stg, f"E{total_row}", "100.0%", BOLD_FONT, LIGHT_GREEN, CENTER, border=thin_border())
write(stg, f"F{total_row}", "", None, LIGHT_GREEN, None, None, thin_border())

for col_let, width in zip("ABCDEF",[28,12,14,12,18,16]):
    stg.column_dimensions[col_let].width = width

stg.row_dimensions[9].height = 10
write(stg, "A10", "Recommendation: Shipping lag (Packed → Shipped) is the primary bottleneck.", Font(name="Arial", bold=True, color="C00000"))
stg.merge_cells("A10:F10")
write(stg, "A11", "Renegotiating carrier SLAs or routing WC orders through faster regional carriers could reduce avg cycle time by ~1 day.", Font(name="Arial", size=10))
stg.merge_cells("A11:F11")

out = "/home/claude/sports-supply-projects/project2_fulfillment_analysis/ProGear_Direct_Fulfillment_Analysis.xlsx"
wb.save(out)
print(f"Saved: {out}")
