import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta
import random

random.seed(99)

DARK_GREEN  = PatternFill("solid", start_color="1E4D2B")
MED_GREEN   = PatternFill("solid", start_color="2E7D32")
LIGHT_GREEN = PatternFill("solid", start_color="D8EFD9")
LIGHT_GRAY  = PatternFill("solid", start_color="F2F2F2")
WHITE_FILL  = PatternFill("solid", start_color="FFFFFF")
RED_FILL    = PatternFill("solid", start_color="FFDDC1")
YELLOW_FILL = PatternFill("solid", start_color="FFFACD")
BLUE_FILL   = PatternFill("solid", start_color="DDEEFF")

HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=11)
TITLE_FONT  = Font(name="Arial", bold=True, size=14)
BLACK_FONT  = Font(name="Arial")
BOLD_FONT   = Font(name="Arial", bold=True)
BLUE_FONT   = Font(name="Arial", color="0000FF")
RED_FONT    = Font(name="Arial", color="C00000", bold=True)

CENTER = Alignment(horizontal="center", vertical="center")
LEFT   = Alignment(horizontal="left",   vertical="center")
RIGHT  = Alignment(horizontal="right",  vertical="center")

def thin_border():
    s = Side(style="thin", color="AAAAAA")
    return Border(left=s, right=s, top=s, bottom=s)

def write(ws, cell, value, font=None, fill=None, align=None, fmt=None, border=None):
    c = ws[cell]
    c.value = value
    if font:   c.font = font
    if fill:   c.fill = fill
    if align:  c.alignment = align
    if fmt:    c.number_format = fmt
    if border: c.border = border

# ── Generate PO data ──────────────────────────────────────────────────────────
suppliers = [
    ("Varsity Sports Mfg.",   "Chicago, IL",    28),
    ("Pacific Athletic Co.",  "Portland, OR",   35),
    ("Champion Source LLC",   "Charlotte, NC",  21),
    ("Apex Equipment Group",  "Columbus, OH",   30),
    ("Elite Sport Supply",    "Dallas, TX",     25),
]

skus = [
    ("PGD-BBALL",  "Basketball (Official Size 7)",      24.50),
    ("PGD-SBALL",  "Soccer Ball (Size 5, Match)",        18.75),
    ("PGD-FBHELM", "Football Helmet (Youth Medium)",    52.00),
    ("PGD-YOGA",   "Yoga Mat (6mm Non-Slip, 6-pack)",   38.00),
    ("PGD-DBELL",  "Adjustable Dumbbell Set (5-50 lb)", 95.00),
    ("PGD-RESIST", "Resistance Band Set (5-piece)",      9.25),
    ("PGD-JUMP",   "Jump Rope (Speed Cable)",             6.50),
    ("PGD-LACR",   "Lacrosse Stick (Adult Composite)",  41.00),
    ("PGD-TENN",   "Tennis Racket (Intermediate)",      33.50),
    ("PGD-GLOVE",  "Baseball Glove (Youth 11.5\")",     28.00),
]

warehouses = ["Northeast (NE)", "Southeast (SE)", "West Coast (WC)"]

statuses_inbound = ["Delivered & Matched", "Delivered & Matched", "Delivered & Matched",
                    "In Transit", "In Transit", "Pending Shipment",
                    "Discrepancy — Qty Short", "Discrepancy — Wrong SKU"]

po_rows = []
receipt_rows = []
start = datetime(2024, 7, 1)

for i in range(1, 31):
    supplier = random.choice(suppliers)
    sku_data = random.choice(skus)
    wh = random.choice(warehouses)
    po_date = start + timedelta(days=random.randint(0, 100))
    qty_ordered = random.choice([50, 75, 100, 150, 200, 250])
    unit_cost = sku_data[2]
    total_cost = qty_ordered * unit_cost
    lead_days = supplier[2] + random.randint(-5, 10)
    expected_delivery = po_date + timedelta(days=lead_days)
    status = random.choice(statuses_inbound)

    if "Delivered" in status or "Discrepancy" in status:
        actual_delivery = expected_delivery + timedelta(days=random.randint(-3, 5))
        days_variance = (actual_delivery - expected_delivery).days
    else:
        actual_delivery = None
        days_variance = None

    if status == "Discrepancy — Qty Short":
        qty_received = qty_ordered - random.randint(5, 25)
    elif status == "Discrepancy — Wrong SKU":
        qty_received = qty_ordered
    elif "Delivered" in status:
        qty_received = qty_ordered
    else:
        qty_received = None

    po_rows.append({
        "PO Number":        f"PO-2024-{1000+i}",
        "PO Date":          po_date,
        "Supplier":         supplier[0],
        "Supplier Location":supplier[1],
        "SKU":              sku_data[0],
        "Product":          sku_data[1],
        "Destination WH":   wh,
        "Qty Ordered":      qty_ordered,
        "Unit Cost ($)":    unit_cost,
        "Total PO Value ($)":total_cost,
        "Expected Delivery":expected_delivery,
        "Actual Delivery":  actual_delivery,
        "Days vs Expected": days_variance,
        "Qty Received":     qty_received,
        "Status":           status,
        "Matched to Receipt": "Yes" if "Matched" in status else ("Pending" if qty_received is None else "No — Review Required"),
    })

    if qty_received is not None:
        receipt_rows.append({
            "Receipt #":        f"REC-2024-{2000+i}",
            "PO Number":        f"PO-2024-{1000+i}",
            "Receipt Date":     actual_delivery,
            "Warehouse":        wh,
            "SKU":              sku_data[0],
            "Product":          sku_data[1],
            "Qty Expected":     qty_ordered,
            "Qty Received":     qty_received,
            "Variance":         qty_received - qty_ordered,
            "Unit Cost ($)":    unit_cost,
            "Total Value ($)":  qty_received * unit_cost,
            "Match Status":     "✅ Matched" if status == "Delivered & Matched" else "⚠️ Discrepancy",
            "Notes":            "" if status == "Delivered & Matched" else (
                f"Short by {qty_ordered - qty_received} units — follow up with {supplier[0]}" if status == "Discrepancy — Qty Short"
                else f"Wrong SKU delivered — coordinate return with {supplier[0]}"
            ),
        })

def build():
    wb = openpyxl.Workbook()

    # ── Sheet 1: PO Dashboard ─────────────────────────────────────────────
    dash = wb.active
    dash.title = "PO Dashboard"
    dash.sheet_view.showGridLines = False

    write(dash, "A1", "ProGear Direct — Purchase Order & Inbound Shipment Tracker", TITLE_FONT)
    dash.merge_cells("A1:P1"); dash.row_dimensions[1].height = 30
    write(dash, "A2", "Period: Jul – Oct 2024  |  30 POs tracked  |  Suppliers: 5  |  Warehouses: NE · SE · WC",
          Font(name="Arial", italic=True, color="555555"))
    dash.merge_cells("A2:P2"); dash.row_dimensions[3].height = 10

    # KPI cards
    total_pos     = len(po_rows)
    total_value   = sum(r["Total PO Value ($)"] for r in po_rows)
    delivered     = sum(1 for r in po_rows if "Delivered" in r["Status"])
    in_transit    = sum(1 for r in po_rows if r["Status"] == "In Transit")
    discrepancies = sum(1 for r in po_rows if "Discrepancy" in r["Status"])
    pending       = sum(1 for r in po_rows if r["Status"] == "Pending Shipment")

    kpis = [
        ("Total POs Issued",     str(total_pos),                     "A"),
        ("Total PO Value",       f"${total_value:,.0f}",             "C"),
        ("Delivered & Matched",  str(delivered),                     "E"),
        ("In Transit",           str(in_transit),                    "G"),
        ("Discrepancies",        str(discrepancies),                 "I"),
    ]
    for label, value, col_let in kpis:
        col = ord(col_let) - 64
        col2 = col + 1
        dash.merge_cells(f"{col_let}4:{get_column_letter(col2)}4")
        dash.merge_cells(f"{col_let}5:{get_column_letter(col2)}5")
        c_lbl = dash.cell(row=4, column=col, value=label)
        c_lbl.font = Font(name="Arial", size=9, color="555555")
        c_lbl.fill = LIGHT_GREEN; c_lbl.alignment = CENTER; c_lbl.border = thin_border()
        dash.cell(row=4, column=col2).fill = LIGHT_GREEN; dash.cell(row=4, column=col2).border = thin_border()
        is_alert = label == "Discrepancies" and discrepancies > 0
        c_val = dash.cell(row=5, column=col, value=value)
        c_val.font = Font(name="Arial", bold=True, size=16, color="C00000" if is_alert else "000000")
        c_val.fill = RED_FILL if is_alert else WHITE_FILL
        c_val.alignment = CENTER; c_val.border = thin_border()
        dash.cell(row=5, column=col2).fill = RED_FILL if is_alert else WHITE_FILL
        dash.cell(row=5, column=col2).border = thin_border()

    dash.row_dimensions[4].height = 18
    dash.row_dimensions[5].height = 28
    dash.row_dimensions[6].height = 10

    for col_let, width in zip("ABCDEFGHIJKLMNOP",
                               [14,14,14,14,14,14,14,14,14,6,6,6,6,6,6,6]):
        dash.column_dimensions[col_let].width = width

    # Supplier performance table
    write(dash, "A7", "Supplier On-Time & Accuracy Performance", BOLD_FONT)
    sup_hdrs = ["Supplier", "Location", "POs Issued", "Delivered", "Avg Days vs Expected", "Discrepancies", "On-Time Rate"]
    for col, h in enumerate(sup_hdrs, 1):
        c = dash.cell(row=8, column=col, value=h)
        c.font = HEADER_FONT; c.fill = MED_GREEN; c.alignment = CENTER; c.border = thin_border()

    from collections import defaultdict
    sup_stats = defaultdict(lambda: {"pos":0,"delivered":0,"day_vars":[],"discrepancies":0,"location":""})
    for r in po_rows:
        s = r["Supplier"]
        sup_stats[s]["pos"] += 1
        sup_stats[s]["location"] = r["Supplier Location"]
        if "Delivered" in r["Status"] or "Discrepancy" in r["Status"]:
            sup_stats[s]["delivered"] += 1
            if r["Days vs Expected"] is not None:
                sup_stats[s]["day_vars"].append(r["Days vs Expected"])
        if "Discrepancy" in r["Status"]:
            sup_stats[s]["discrepancies"] += 1

    for row_idx, (sup, stats) in enumerate(sorted(sup_stats.items()), 9):
        fill = LIGHT_GRAY if row_idx % 2 == 0 else WHITE_FILL
        avg_var = round(sum(stats["day_vars"]) / len(stats["day_vars"]), 1) if stats["day_vars"] else 0
        on_time = round((stats["delivered"] - stats["discrepancies"]) / max(stats["pos"], 1) * 100, 1)
        disc = stats["discrepancies"]
        vals = [sup, stats["location"], stats["pos"], stats["delivered"], avg_var, disc, f"{on_time}%"]
        for col, val in enumerate(vals, 1):
            c = dash.cell(row=row_idx, column=col, value=val)
            c.font = RED_FONT if (col == 6 and disc > 0) else BLACK_FONT
            c.fill = RED_FILL if (col == 6 and disc > 0) else fill
            c.alignment = LEFT if col in (1,2) else CENTER
            c.border = thin_border()
            if col == 5: c.number_format = '+0.0;-0.0;0.0'

    dash.row_dimensions[14].height = 10

    # Status summary
    write(dash, "A15", "PO Status Summary", BOLD_FONT)
    status_hdrs = ["Status", "Count", "% of Total"]
    for col, h in enumerate(status_hdrs, 1):
        c = dash.cell(row=16, column=col, value=h)
        c.font = HEADER_FONT; c.fill = DARK_GREEN; c.alignment = CENTER; c.border = thin_border()

    status_counts = {}
    for r in po_rows:
        status_counts[r["Status"]] = status_counts.get(r["Status"], 0) + 1

    status_fills = {
        "Delivered & Matched":    PatternFill("solid", start_color="E2EFDA"),
        "In Transit":             PatternFill("solid", start_color="DDEEFF"),
        "Pending Shipment":       LIGHT_GRAY,
        "Discrepancy — Qty Short":RED_FILL,
        "Discrepancy — Wrong SKU":RED_FILL,
    }
    for row_idx, (status, count) in enumerate(sorted(status_counts.items()), 17):
        fill = status_fills.get(status, WHITE_FILL)
        pct = round(count / total_pos * 100, 1)
        for col, val in enumerate([status, count, f"{pct}%"], 1):
            c = dash.cell(row=row_idx, column=col, value=val)
            c.font = BLACK_FONT; c.fill = fill
            c.alignment = LEFT if col == 1 else CENTER; c.border = thin_border()

    # ── Sheet 2: PO Log ───────────────────────────────────────────────────
    po_sheet = wb.create_sheet("PO Log")
    po_sheet.sheet_view.showGridLines = True

    po_headers = list(po_rows[0].keys())
    for col, h in enumerate(po_headers, 1):
        c = po_sheet.cell(row=1, column=col, value=h)
        c.font = HEADER_FONT; c.fill = DARK_GREEN; c.alignment = CENTER; c.border = thin_border()

    po_widths = [16,13,22,16,12,30,18,13,13,16,17,15,16,14,24,20]
    for i, w in enumerate(po_widths[:len(po_headers)], 1):
        po_sheet.column_dimensions[get_column_letter(i)].width = w

    status_row_fills = {
        "Delivered & Matched":    PatternFill("solid", start_color="E2EFDA"),
        "In Transit":             PatternFill("solid", start_color="DDEEFF"),
        "Pending Shipment":       WHITE_FILL,
        "Discrepancy — Qty Short":RED_FILL,
        "Discrepancy — Wrong SKU":RED_FILL,
    }

    for r, row_data in enumerate(po_rows, 2):
        row_fill = status_row_fills.get(row_data["Status"], WHITE_FILL)
        for col, (key, val) in enumerate(row_data.items(), 1):
            c = po_sheet.cell(row=r, column=col, value=val)
            c.font = BLACK_FONT; c.fill = row_fill
            c.alignment = LEFT if key in ("Product","Supplier") else CENTER
            c.border = thin_border()
            if key in ("PO Date","Expected Delivery","Actual Delivery") and isinstance(val, datetime):
                c.number_format = "MM/DD/YYYY"
            if key in ("Unit Cost ($)","Total PO Value ($)"): c.number_format = '$#,##0.00'
            if key == "Days vs Expected" and isinstance(val, int):
                c.number_format = '+0;-0;0'
                c.font = Font(name="Arial", color="C00000" if val > 3 else "000000")

    # ── Sheet 3: Receipt Log ──────────────────────────────────────────────
    rec_sheet = wb.create_sheet("Receipt Log")
    rec_sheet.sheet_view.showGridLines = True

    rec_headers = list(receipt_rows[0].keys())
    for col, h in enumerate(rec_headers, 1):
        c = rec_sheet.cell(row=1, column=col, value=h)
        c.font = HEADER_FONT; c.fill = MED_GREEN; c.alignment = CENTER; c.border = thin_border()

    rec_widths = [16,16,13,18,12,30,14,14,12,13,14,16,38]
    for i, w in enumerate(rec_widths[:len(rec_headers)], 1):
        rec_sheet.column_dimensions[get_column_letter(i)].width = w

    for r, row_data in enumerate(receipt_rows, 2):
        is_disc = row_data["Match Status"] != "✅ Matched"
        row_fill = RED_FILL if is_disc else (LIGHT_GRAY if r % 2 == 0 else WHITE_FILL)
        for col, (key, val) in enumerate(row_data.items(), 1):
            c = rec_sheet.cell(row=r, column=col, value=val)
            c.font = BLACK_FONT; c.fill = row_fill
            c.alignment = LEFT if key in ("Product","Notes") else CENTER
            c.border = thin_border()
            if key == "Receipt Date" and isinstance(val, datetime):
                c.number_format = "MM/DD/YYYY"
            if key in ("Unit Cost ($)","Total Value ($)"): c.number_format = '$#,##0.00'
            if key == "Variance":
                c.number_format = '+0;-0;0'
                if isinstance(val, int) and val != 0:
                    c.font = Font(name="Arial", color="C00000", bold=True)

    # ── Sheet 4: Discrepancy Report ───────────────────────────────────────
    disc_sheet = wb.create_sheet("Discrepancy Report")
    disc_sheet.sheet_view.showGridLines = False

    write(disc_sheet, "A1", "Discrepancy Report — Items Requiring Follow-Up", TITLE_FONT)
    disc_sheet.merge_cells("A1:M1"); disc_sheet.row_dimensions[1].height = 28
    write(disc_sheet, "A2", "Review each discrepancy below and coordinate resolution with the supplier. Update status once resolved.",
          Font(name="Arial", italic=True, color="555555"))
    disc_sheet.merge_cells("A2:M2"); disc_sheet.row_dimensions[3].height = 8

    disc_hdrs = ["PO Number","Receipt #","Supplier","SKU","Product","Destination WH",
                 "Qty Ordered","Qty Received","Variance","Issue Type","Supplier Contact Action","Resolution Status","Notes"]
    for col, h in enumerate(disc_hdrs, 1):
        c = disc_sheet.cell(row=4, column=col, value=h)
        c.font = HEADER_FONT; c.fill = PatternFill("solid", start_color="8B0000"); c.alignment = CENTER; c.border = thin_border()

    disc_widths = [16,14,22,12,28,18,13,13,10,22,26,18,32]
    for i, w in enumerate(disc_widths, 1):
        disc_sheet.column_dimensions[get_column_letter(i)].width = w

    disc_po_rows   = [r for r in po_rows    if "Discrepancy" in r["Status"]]
    disc_rec_rows  = {r["PO Number"]: r for r in receipt_rows if r["Match Status"] == "⚠️ Discrepancy"}

    for r, po in enumerate(disc_po_rows, 5):
        rec = disc_rec_rows.get(po["PO Number"], {})
        issue = "Quantity short — fewer units received than ordered" if "Qty Short" in po["Status"] else "Wrong SKU delivered — product mismatch"
        action = "Request credit or replacement shipment" if "Qty Short" in po["Status"] else "Arrange return shipment; reorder correct SKU"
        variance = (rec.get("Qty Received", po["Qty Ordered"]) - po["Qty Ordered"])
        vals = [
            po["PO Number"], rec.get("Receipt #","—"), po["Supplier"],
            po["SKU"], po["Product"], po["Destination WH"],
            po["Qty Ordered"], rec.get("Qty Received","—"),
            variance if rec else "—",
            issue, action, "Open — Pending Supplier Response", rec.get("Notes","")
        ]
        for col, val in enumerate(vals, 1):
            c = disc_sheet.cell(row=r, column=col, value=val)
            c.font = BLACK_FONT; c.fill = RED_FILL if r % 2 == 0 else PatternFill("solid", start_color="FFE8E0")
            c.alignment = LEFT if col in (3,5,10,11,13) else CENTER
            c.border = thin_border()
            if col == 9 and isinstance(val, int):
                c.number_format = '+0;-0;0'
                c.font = Font(name="Arial", color="C00000", bold=True)

    out = "/home/claude/sports-supply-projects/project3_po_tracker/ProGear_Direct_PO_Inbound_Tracker.xlsx"
    wb.save(out)
    print(f"Saved: {out}")

build()
