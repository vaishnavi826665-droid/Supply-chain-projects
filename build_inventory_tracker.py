import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BLUE_FONT   = Font(name="Arial", color="0000FF")
BLACK_FONT  = Font(name="Arial", color="000000")
BOLD_FONT   = Font(name="Arial", bold=True)
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=11)
TITLE_FONT  = Font(name="Arial", bold=True, size=14)

DARK_GREEN  = PatternFill("solid", start_color="1E4D2B")
MED_GREEN   = PatternFill("solid", start_color="2E7D32")
LIGHT_GREEN = PatternFill("solid", start_color="D8EFD9")
LIGHT_GRAY  = PatternFill("solid", start_color="F2F2F2")
WHITE_FILL  = PatternFill("solid", start_color="FFFFFF")

CENTER = Alignment(horizontal="center", vertical="center")
LEFT   = Alignment(horizontal="left",   vertical="center")

def thin_border():
    s = Side(style="thin", color="AAAAAA")
    return Border(left=s, right=s, top=s, bottom=s)

def write(ws, cell, value, font=None, fill=None, align=None, fmt=None, border=None):
    c = ws[cell]
    c.value = value
    if font:   c.font = font
    if fill:   c.fill = fill
    if align:  c.alignment = align
    if fmt:    c.number_format = fmt
    if border: c.border = border

def build():
    wb = openpyxl.Workbook()

    # ── Sheet 1: Dashboard ────────────────────────────────────────────────
    dash = wb.active
    dash.title = "Dashboard"
    dash.sheet_view.showGridLines = False

    write(dash, "A1", "ProGear Direct — Multi-Warehouse Inventory Rebalancer", TITLE_FONT)
    dash.merge_cells("A1:L1")
    dash["A1"].alignment = LEFT
    write(dash, "A2", "Monitoring Period: Sep – Nov  |  Warehouses: Northeast (NE) · Southeast (SE) · West Coast (WC)", Font(name="Arial", italic=True, color="555555"))
    dash.merge_cells("A2:L2")
    dash.row_dimensions[1].height = 30
    dash.row_dimensions[2].height = 18
    dash.row_dimensions[3].height = 8

    headers = ["SKU", "Product", "Min Stock", "NE On Hand", "SE On Hand", "WC On Hand",
               "Total On Hand", "Total Demand (Wk)", "Weeks of Supply", "NE Status", "SE Status", "WC Status"]
    for col, h in enumerate(headers, 1):
        cell = dash.cell(row=4, column=col, value=h)
        cell.font = HEADER_FONT; cell.fill = DARK_GREEN
        cell.alignment = CENTER; cell.border = thin_border()

    col_widths = [12, 32, 11, 12, 12, 12, 14, 16, 15, 12, 12, 12]
    for i, w in enumerate(col_widths, 1):
        dash.column_dimensions[get_column_letter(i)].width = w

    # SKU, Product, Min Stock, NE, SE, WC, Weekly Demand
    skus = [
        ("PGD-BBALL",  "Basketball (Official Size 7)",        30, 48, 12, 35,  95),
        ("PGD-SBALL",  "Soccer Ball (Size 5, Match)",         40, 15, 55, 20, 110),
        ("PGD-FBHELM", "Football Helmet (Youth Medium)",      20,  8, 30, 14,  50),
        ("PGD-YOGA",   "Yoga Mat (6mm Non-Slip, 6-pack)",     25, 60, 18, 10,  75),
        ("PGD-DBELL",  "Adjustable Dumbbell Set (5-50 lb)",   15,  6, 28, 22,  45),
        ("PGD-RESIST", "Resistance Band Set (5-piece)",       50,130, 40, 80, 160),
        ("PGD-JUMP",   "Jump Rope (Speed Cable)",             60, 20, 90, 45, 140),
        ("PGD-LACR",   "Lacrosse Stick (Adult Composite)",    20,  5, 14, 32,  48),
        ("PGD-TENN",   "Tennis Racket (Intermediate)",        25, 55, 10, 18,  65),
        ("PGD-GLOVE",  "Baseball Glove (Youth 11.5\")",       30, 12, 45, 26,  70),
    ]

    for r, (sku, product, min_s, ne, se, wc, demand) in enumerate(skus, 5):
        fill = LIGHT_GRAY if r % 2 == 0 else WHITE_FILL
        for col, val in enumerate([sku, product, min_s, ne, se, wc], 1):
            c = dash.cell(row=r, column=col, value=val)
            c.font = BLUE_FONT if col == 3 else BLACK_FONT
            c.fill = fill
            c.alignment = LEFT if col == 2 else CENTER
            c.border = thin_border()
            if col in (3,4,5,6): c.number_format = "#,##0"

        for col, formula in [
            (7, f"=D{r}+E{r}+F{r}"),
            (9, f"=IFERROR(G{r}/H{r},\"-\")"),
        ]:
            c = dash.cell(row=r, column=col, value=formula)
            c.font = BLACK_FONT; c.fill = fill; c.alignment = CENTER; c.border = thin_border()
            c.number_format = "#,##0" if col == 7 else "0.0"

        c = dash.cell(row=r, column=8, value=demand)
        c.font = BLUE_FONT; c.fill = fill; c.alignment = CENTER; c.border = thin_border(); c.number_format = "#,##0"

        for col_idx, wh_col in [(10,"D"),(11,"E"),(12,"F")]:
            formula = (f'=IF({wh_col}{r}<C{r}*0.5,"🔴 CRITICAL",'
                       f'IF({wh_col}{r}<C{r},"⚠️ LOW","✅ OK"))')
            c = dash.cell(row=r, column=col_idx, value=formula)
            c.font = BLACK_FONT; c.fill = fill; c.alignment = CENTER; c.border = thin_border()

    last_data_row = 4 + len(skus)
    dash.row_dimensions[last_data_row + 1].height = 4

    tr = last_data_row + 2
    write(dash, f"A{tr}", "TOTALS", BOLD_FONT, LIGHT_GREEN, CENTER, border=thin_border())
    dash.merge_cells(f"A{tr}:C{tr}")
    for col in range(4, 10):
        col_l = get_column_letter(col)
        val = f"=SUM({col_l}5:{col_l}{last_data_row})" if col != 9 else f"=IFERROR(G{tr}/H{tr},\"-\")"
        c = dash.cell(row=tr, column=col, value=val)
        c.font = BOLD_FONT; c.fill = LIGHT_GREEN; c.alignment = CENTER; c.border = thin_border()
        c.number_format = "#,##0" if col != 9 else "0.0"
    for col in range(10, 13):
        dash.cell(row=tr, column=col, value="").fill = LIGHT_GREEN
        dash.cell(row=tr, column=col).border = thin_border()

    leg = tr + 2
    write(dash, f"A{leg}",   "Legend", BOLD_FONT)
    write(dash, f"A{leg+1}", "✅ OK",  BLACK_FONT)
    write(dash, f"B{leg+1}", "On hand ≥ min stock", Font(name="Arial", size=9, color="555555"))
    write(dash, f"A{leg+2}", "⚠️ LOW", BLACK_FONT)
    write(dash, f"B{leg+2}", "On hand < min stock", Font(name="Arial", size=9, color="555555"))
    write(dash, f"A{leg+3}", "🔴 CRITICAL", BLACK_FONT)
    write(dash, f"B{leg+3}", "On hand < 50% of min — transfer recommended", Font(name="Arial", size=9, color="555555"))
    write(dash, f"A{leg+5}", "Blue cells = editable input values", Font(name="Arial", italic=True, color="0000FF", size=9))

    # ── Sheet 2: Transfer Recommendations ────────────────────────────────
    tr_sheet = wb.create_sheet("Transfer Recommendations")
    tr_sheet.sheet_view.showGridLines = False
    write(tr_sheet, "A1", "Transfer Recommendations — Back-to-School & Fall Season", TITLE_FONT)
    tr_sheet.merge_cells("A1:H1"); tr_sheet.row_dimensions[1].height = 28
    write(tr_sheet, "A2", "Update on-hand values in Dashboard tab to refresh status flags, then revise quantities below.", Font(name="Arial", italic=True, color="555555"))
    tr_sheet.merge_cells("A2:H2"); tr_sheet.row_dimensions[3].height = 6

    tr_headers = ["SKU", "Product", "From Warehouse", "To Warehouse", "Qty to Transfer", "Priority", "Est. Transit Days", "Notes"]
    for col, h in enumerate(tr_headers, 1):
        c = tr_sheet.cell(row=4, column=col, value=h)
        c.font = HEADER_FONT; c.fill = MED_GREEN; c.alignment = CENTER; c.border = thin_border()
    for i, w in enumerate([12,30,18,18,15,12,16,34], 1):
        tr_sheet.column_dimensions[get_column_letter(i)].width = w

    transfers = [
        ("PGD-BBALL",  "Basketball (Official Size 7)",     "Northeast (NE)", "Southeast (SE)", 20, "HIGH",   2, "SE critically low; NE has surplus vs demand"),
        ("PGD-DBELL",  "Adjustable Dumbbell Set",          "Southeast (SE)", "Northeast (NE)", 12, "HIGH",   2, "NE below 50% min; SE well-stocked"),
        ("PGD-LACR",   "Lacrosse Stick (Adult Composite)", "West Coast (WC)","Northeast (NE)",  8, "HIGH",   4, "NE critically low — only 5 units on hand"),
        ("PGD-TENN",   "Tennis Racket (Intermediate)",     "Northeast (NE)", "Southeast (SE)", 18, "MEDIUM", 2, "SE approaching min threshold"),
        ("PGD-GLOVE",  "Baseball Glove (Youth 11.5\")",    "Southeast (SE)", "West Coast (WC)", 10,"MEDIUM", 4, "WC below min; SE has comfortable surplus"),
        ("PGD-JUMP",   "Jump Rope (Speed Cable)",          "Southeast (SE)", "Northeast (NE)", 25, "LOW",    2, "NE slightly below min; SE heavily overstocked"),
    ]
    pri_fills = {"HIGH": PatternFill("solid", start_color="FFDDC1"), "MEDIUM": PatternFill("solid", start_color="FFFACD"), "LOW": PatternFill("solid", start_color="E2EFDA")}
    for r, row in enumerate(transfers, 5):
        base_fill = LIGHT_GRAY if r % 2 == 0 else WHITE_FILL
        for col, val in enumerate(row, 1):
            c = tr_sheet.cell(row=r, column=col, value=val)
            c.font = BLACK_FONT
            c.fill = pri_fills.get(val, base_fill) if col == 6 else base_fill
            c.alignment = LEFT if col in (2, 8) else CENTER
            c.border = thin_border()
            if col == 5: c.number_format = "#,##0"

    # ── Sheet 3: Raw Data ─────────────────────────────────────────────────
    raw = wb.create_sheet("Raw Data")
    write(raw, "A1", "Raw Inventory Data — Edit here; Dashboard pulls from this sheet", Font(name="Arial", bold=True, italic=True, color="555555"))
    raw.merge_cells("A1:G1")
    raw_headers = ["SKU", "Product", "Min Stock", "NE On Hand", "SE On Hand", "WC On Hand", "Weekly Demand"]
    for col, h in enumerate(raw_headers, 1):
        c = raw.cell(row=2, column=col, value=h)
        c.font = HEADER_FONT; c.fill = DARK_GREEN; c.alignment = CENTER; c.border = thin_border()
    raw.column_dimensions["A"].width = 12
    raw.column_dimensions["B"].width = 32
    for i in range(3, 8):
        raw.column_dimensions[get_column_letter(i)].width = 14
    for r, (sku, product, min_s, ne, se, wc, demand) in enumerate(skus, 3):
        fill = LIGHT_GRAY if r % 2 == 0 else WHITE_FILL
        for col, val in enumerate([sku, product, min_s, ne, se, wc, demand], 1):
            c = raw.cell(row=r, column=col, value=val)
            c.font = BLUE_FONT if col > 2 else BLACK_FONT
            c.fill = fill; c.alignment = LEFT if col == 2 else CENTER
            c.border = thin_border()
            if col > 2: c.number_format = "#,##0"

    out = "/home/claude/sports-supply-projects/project1_inventory_rebalancer/ProGear_Direct_Inventory_Rebalancer.xlsx"
    wb.save(out)
    print(f"Saved: {out}")

build()
